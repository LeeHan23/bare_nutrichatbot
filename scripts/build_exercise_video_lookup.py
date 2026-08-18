"""
Flatten "Exercise Video Intensity.xlsx"'s Ex Intervention Tag tab into
data/exercise_video_lookup.json — one record per (exercise, intensity tier).

Several columns (Type, Exercise Title, Explanation, Body Focus, Suitable
For, YouTube Link, Video Duration) are only populated on the first sub-row
of each exercise's group in the source sheet (merged-cell export); this
script forward-fills them. Intensity Tier / Allowed Level / loop counts /
Total Duration are already present on every row.

Re-run (full overwrite) whenever the client resends the sheet:

    python scripts/build_exercise_video_lookup.py "Exercise Video Intensity.xlsx"
"""
import argparse
import json
import os
import sys

_PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, _PROJECT_ROOT)

import openpyxl

OUT_PATH = os.path.join(_PROJECT_ROOT, "data", "exercise_video_lookup.json")

# Columns that are only populated on the first sub-row of each exercise
# group and need forward-filling.
GROUPED_COLUMNS = [
    "Type", "Exercise Title", "Video Exercise Explanation (Based on Video Content)",
    "Body Focus", "Suitable For / Purpose", "YouTube Link", "Video Duration (min:sec)",
]


def build(workbook_path: str) -> list[dict]:
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    ws = wb["Ex Intervention Tag"]

    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row))
    header_row_idx = None
    for i, row in enumerate(rows):
        if row[1].value == "Type":  # column B
            header_row_idx = i
            break
    if header_row_idx is None:
        raise SystemExit("Could not find header row (column B == 'Type')")

    header = [c.value for c in rows[header_row_idx]]
    records = []
    carry = {}
    for row in rows[header_row_idx + 1:]:
        values = dict(zip(header, [c.value for c in row]))
        if not any(v is not None and str(v).strip() for v in values.values()):
            continue

        for col in GROUPED_COLUMNS:
            if values.get(col) not in (None, ""):
                carry[col] = values[col]
            else:
                values[col] = carry.get(col)

        video_duration = values.get("Video Duration (min:sec)")
        records.append({
            "type": values.get("Type"),
            "exercise_title": values.get("Exercise Title"),
            "explanation": values.get("Video Exercise Explanation (Based on Video Content)"),
            "intensity_tier": values.get("Intensity Tier"),
            "min_loop": values.get("Min Loop"),
            "max_loop": values.get("Max Loop"),
            "total_duration": values.get("Total Duration"),
            "allowed_level": values.get("Allowed Level"),
            "body_focus": values.get("Body Focus"),
            "suitable_for": values.get("Suitable For / Purpose"),
            "youtube_link": values.get("YouTube Link"),
            "video_duration": str(video_duration) if video_duration is not None else None,
        })

    return records


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", help='Path to "Exercise Video Intensity.xlsx"')
    args = parser.parse_args()

    records = build(args.workbook)
    with_link = [r for r in records if r["youtube_link"]]
    print(f"Flattened {len(records)} (exercise, intensity) rows, {len(with_link)} with a YouTube link")

    os.makedirs(os.path.dirname(OUT_PATH), exist_ok=True)
    with open(OUT_PATH, "w", encoding="utf-8") as f:
        json.dump(records, f, indent=2, ensure_ascii=False, default=str)
    print(f"Wrote {OUT_PATH}")


if __name__ == "__main__":
    main()
