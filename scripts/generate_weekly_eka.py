"""
generate_weekly_eka.py — Weekly Exercise / Knowledge / Activity content generator.

Generates three content types per condition group on a 4-week rotating schedule:
  E — Exercise: structured sessions with warm-up, main activity, cool-down, level modifications
  K — Knowledge: 6 educational points on a health topic
  A — Activity: daily/weekly behavioural task with micro-actions

Usage:
    # Generate all E/K/A for current ISO week
    python scripts/generate_weekly_eka.py

    # Generate for a specific week number
    python scripts/generate_weekly_eka.py --week 22

    # Filter to one group or type
    python scripts/generate_weekly_eka.py --group CKD --type K

    # Dry run — show what would be generated, no LLM call
    python scripts/generate_weekly_eka.py --dry-run

    # Force overwrite existing DB rows
    python scripts/generate_weekly_eka.py --force
"""
import argparse, json, os, sys
from datetime import datetime, date

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

# ---------------------------------------------------------------------------
# 4-week rotating topic library per condition group × content type
# ---------------------------------------------------------------------------
# Structure: group_slug → content_type → list of 4 (week_slot, topic_slug, title, rag_query, prompt_topic)
# week_slot 1-4 cycles via: rotation = ((iso_week - 1) % 4) + 1

_TEMPLATES = {
    "T2DM": {
        "condition_tags": ["Type 2 Diabetes"],
        "E": [
            (1, "walking_foundation",
             "Week {w} — Walking Foundation for Blood Sugar",
             "walking aerobic exercise type 2 diabetes blood glucose safety",
             "a 20-25 min beginner walking programme for T2DM, blood glucose checks before/after exercise, "
             "safe intensity using talk test, what to do if glucose is too low before starting"),
            (2, "resistance_basics",
             "Week {w} — Resistance Training for Insulin Sensitivity",
             "resistance strength training type 2 diabetes insulin sensitivity bodyweight bands",
             "simple resistance exercises (bodyweight or bands) 2×/week to improve insulin sensitivity for T2DM in Malaysia, "
             "including progressions from seated to standing movements"),
            (3, "mixed_training",
             "Week {w} — Combining Cardio and Strength",
             "combined aerobic resistance exercise diabetes HbA1c blood sugar gym-free",
             "a 40-min combined session: 20 min aerobic + 20 min strength, gym-free, safe blood glucose management"),
            (4, "flexibility_balance",
             "Week {w} — Flexibility and Balance for Diabetes",
             "stretching flexibility yoga diabetes peripheral neuropathy balance foot care",
             "gentle stretching and balance exercises addressing peripheral neuropathy risk in T2DM, "
             "with foot inspection reminder as part of cool-down"),
        ],
        "K": [
            (1, "understanding_t2dm",
             "Week {w} — Understanding Type 2 Diabetes",
             "type 2 diabetes pathophysiology insulin resistance beta cell progression Malaysia",
             "what T2DM is, why insulin resistance develops, and how Malaysian diet and lifestyle patterns drive progression"),
            (2, "blood_sugar_food",
             "Week {w} — How Food Affects Blood Sugar",
             "glycaemic index glycaemic load carbohydrate blood glucose response Malaysian food",
             "how GI and GL of common Malaysian foods (nasi, roti, mee, kuih) affect blood sugar and how to reduce spikes"),
            (3, "exercise_insulin",
             "Week {w} — Why Exercise Is Medicine for Diabetes",
             "exercise diabetes HbA1c insulin sensitivity glucose uptake GLUT4 benefit",
             "the science of GLUT4-mediated glucose uptake during exercise, how it lowers blood sugar independent of insulin"),
            (4, "monitoring_targets",
             "Week {w} — Monitoring Your Blood Sugar",
             "blood glucose self-monitoring HbA1c fasting postprandial targets home glucometer",
             "how to use a glucometer, what fasting and post-meal glucose targets mean, when to call the clinic"),
        ],
        "A": [
            (1, "step_tracking",
             "Week {w} — Daily Step Tracking Challenge",
             "daily steps walking diabetes blood sugar benefit 5000 10000",
             "setting a daily step goal starting at 5000 steps/day, how to track, and the blood sugar benefit"),
            (2, "glucose_diary",
             "Week {w} — Blood Glucose and Meal Diary",
             "blood glucose diary food log self-monitoring diabetes patterns",
             "a simple food + blood glucose log to identify personal trigger foods and meal-timing patterns"),
            (3, "meal_timing",
             "Week {w} — Meal Timing Tracker",
             "meal timing regularity carbohydrate distribution diabetes blood sugar spikes",
             "tracking meal times and carb portions across the day to smooth blood sugar fluctuations"),
            (4, "monthly_review",
             "Week {w} — Monthly Progress Check-In",
             "diabetes self-assessment HbA1c goals progress lifestyle review",
             "4-week review of steps, glucose logs, and lifestyle changes; setting the next month's goals"),
        ],
    },
    "HTN": {
        "condition_tags": ["Hypertension"],
        "E": [
            (1, "gentle_walking",
             "Week {w} — Gentle Walking for Blood Pressure",
             "walking aerobic exercise hypertension blood pressure reduction safe",
             "a 20-min gentle walking programme for hypertension, RPE scale for effort, "
             "BP monitoring before and after, hydration in Malaysian heat"),
            (2, "low_impact_cardio",
             "Week {w} — Low-Impact Cardio Routine",
             "low impact exercise hypertension swimming cycling light aerobic blood pressure",
             "low-impact cardio options available in Malaysia (pool walking, cycling, light aerobics) safe for hypertension, "
             "how to build to 150 min/week gradually"),
            (3, "breathing_exercises",
             "Week {w} — Breathing and Relaxation Exercises",
             "deep breathing diaphragmatic breathing blood pressure reduction hypertension stress",
             "slow deep breathing (4-7-8 technique, diaphragmatic breathing) to acutely lower BP and reduce sympathetic tone"),
            (4, "gentle_yoga",
             "Week {w} — Gentle Yoga and Stretching",
             "yoga stretching hypertension blood pressure parasympathetic relaxation",
             "gentle yoga poses for hypertension: which inversions to avoid, best poses for BP reduction"),
        ],
        "K": [
            (1, "understanding_htn",
             "Week {w} — Understanding Hypertension",
             "hypertension pathophysiology blood pressure stages causes risk Malaysia prevalence",
             "what BP numbers mean (systolic/diastolic), how hypertension damages vessels, why 'silent killer' matters in Malaysia"),
            (2, "sodium_bp",
             "Week {w} — Sodium, Potassium, and Blood Pressure",
             "sodium potassium blood pressure DASH diet hypertension reduction mechanism",
             "the mechanism of sodium-BP link, potassium as counter-balance, and examples from Malaysian cooking (kicap, belacan, monosodium glutamate)"),
            (3, "stress_bp",
             "Week {w} — Stress and Blood Pressure",
             "stress cortisol adrenaline blood pressure hypertension sympathetic nervous system",
             "how work stress, financial stress, and social pressures raise BP via cortisol and adrenaline, "
             "and evidence-based coping strategies"),
            (4, "bp_monitoring",
             "Week {w} — Understanding Your Blood Pressure Numbers",
             "blood pressure measurement technique home monitoring systolic diastolic target hypertension",
             "correct BP measurement technique at home, white-coat effect, treatment targets by guideline"),
        ],
        "A": [
            (1, "bp_tracking",
             "Week {w} — Twice-Daily Blood Pressure Log",
             "blood pressure home monitoring log morning evening hypertension tracking",
             "setting up a twice-daily BP logging habit (morning before meds, evening) with simple chart"),
            (2, "sodium_audit",
             "Week {w} — Sodium Audit Challenge",
             "sodium food label reading hypertension high sodium Malaysian foods audit",
             "auditing 5 commonly used condiments and packaged foods at home for sodium content"),
            (3, "stress_checkin",
             "Week {w} — Daily Stress Check-In",
             "stress monitoring mood BP hypertension daily habit awareness",
             "a 2-min daily stress rating (1-10) + trigger note, to identify BP-driving stressors"),
            (4, "weekly_bp_review",
             "Week {w} — Weekly BP Trend Review",
             "blood pressure trend analysis hypertension progress dietary lifestyle response",
             "reviewing 4 weeks of BP readings: identifying patterns, diet + exercise correlations, progress"),
        ],
    },
    "CKD": {
        "condition_tags": ["Chronic Kidney Disease"],
        "E": [
            (1, "gentle_mobility",
             "Week {w} — Gentle Mobility for Kidney Patients",
             "exercise CKD chronic kidney disease safety gentle mobility fatigue management",
             "very gentle range-of-motion and mobility exercises safe for CKD, energy conservation principles, "
             "when to rest vs. push through fatigue"),
            (2, "seated_exercise",
             "Week {w} — Seated Exercise Routine",
             "seated exercise chair workout CKD chronic kidney disease anaemia fatigue",
             "a 15-min seated exercise circuit for CKD patients with fatigue or mobility issues, "
             "including arm raises, seated marching, ankle rotations"),
            (3, "light_walking",
             "Week {w} — Light Progressive Walking",
             "walking CKD kidney disease anaemia fatigue blood pressure exercise tolerance",
             "a progressive walking programme for CKD starting at 10 min: managing anaemia-related fatigue, "
             "monitoring for oedema and shortness of breath"),
            (4, "balance_safety",
             "Week {w} — Balance and Fall Prevention",
             "balance training fall prevention CKD kidney disease elderly neuropathy",
             "balance exercises to reduce fall risk in CKD (tandem stand, single-leg stand with support), "
             "footwear and home safety tips"),
        ],
        "K": [
            (1, "understanding_ckd",
             "Week {w} — Understanding Chronic Kidney Disease",
             "chronic kidney disease CKD eGFR stages kidneys filtration function Malaysia",
             "what eGFR stages mean (1-5), how kidneys filter waste, and what declining function means for diet in Malaysia"),
            (2, "diet_kidneys",
             "Week {w} — How Diet Affects Your Kidneys",
             "kidney diet protein potassium phosphorus sodium CKD nutrition impact",
             "how protein, potassium, phosphorus, and sodium each stress or protect kidneys, with Malaysian food examples"),
            (3, "fluid_kidneys",
             "Week {w} — Fluid Balance and CKD",
             "fluid restriction CKD oedema fluid overload signs measurement urine output",
             "why fluid restriction matters in CKD, how to measure daily intake, signs of fluid overload (ankle swelling, breathlessness)"),
            (4, "ckd_progression",
             "Week {w} — Slowing CKD Progression",
             "CKD progression slowing blood pressure diabetes control diet eGFR",
             "evidence-based strategies to slow CKD: BP < 130/80, HbA1c < 7%, low-protein diet, RAAS inhibitors"),
        ],
        "A": [
            (1, "fluid_tracking",
             "Week {w} — Daily Fluid Intake Tracker",
             "fluid intake tracking CKD kidney disease restriction daily habit",
             "setting up a daily fluid tracking habit using a marked water bottle, converting common drinks to ml"),
            (2, "daily_weight",
             "Week {w} — Daily Weight Monitoring",
             "weight monitoring fluid retention CKD oedema daily morning weigh-in",
             "daily morning weigh-in to detect fluid retention early: what gain over 2 days should trigger a call to clinic"),
            (3, "kidney_food_diary",
             "Week {w} — Kidney-Safe Food Diary",
             "food diary CKD kidney disease potassium phosphorus sodium tracking identification",
             "food diary focused on 3 key minerals: marking high-K, high-P, high-Na foods daily"),
            (4, "lab_tracker",
             "Week {w} — Understanding Your Lab Values",
             "CKD lab values eGFR creatinine potassium phosphorus bicarbonate monitoring",
             "a patient-friendly explanation of 5 key CKD lab values, what trends mean, and what to report to doctor"),
        ],
    },
    "Cardiac": {
        "condition_tags": ["Ischaemic Heart Disease", "Heart Failure", "Post-CABG"],
        "E": [
            (1, "cardiac_walk_phase1",
             "Week {w} — Phase 1 Cardiac Walking Programme",
             "cardiac rehabilitation phase 1 walking post-CABG heart failure exercise safe",
             "Phase 1 cardiac rehab: 10-15 min flat-ground walking at RPE 9-11 (very light), "
             "pulse monitoring, stop-criteria, temperature safety in Malaysian climate"),
            (2, "cardiac_walk_phase2",
             "Week {w} — Phase 2 Cardiac Walking Progression",
             "cardiac rehab phase 2 walking exercise progression heart disease safe",
             "Progressing to 20-25 min at RPE 11-13 (light-moderate), "
             "interval approach (walk 5, rest 1), Borg scale use, warning signs"),
            (3, "upper_body_gentle",
             "Week {w} — Gentle Upper Body Movements",
             "upper body exercise cardiac patients heart failure seated arm movements circulation",
             "gentle seated upper-body movements (shoulder circles, arm raises < shoulder height) "
             "to improve circulation without raising cardiac demand significantly"),
            (4, "breathing_cardiac",
             "Week {w} — Breathing Exercises for Heart Health",
             "pursed lip breathing diaphragmatic breathing heart failure dyspnoea cardiac",
             "pursed-lip breathing and diaphragmatic breathing to reduce dyspnoea, "
             "improve oxygen efficiency, and calm sympathetic activation in heart failure"),
        ],
        "K": [
            (1, "understanding_heart",
             "Week {w} — Understanding Your Heart Condition",
             "heart disease ischaemic heart failure ejection fraction pathophysiology Malaysia",
             "what IHD or heart failure means in plain language, what ejection fraction is, "
             "and why diet and activity matter for recovery"),
            (2, "cholesterol_heart",
             "Week {w} — Cholesterol and Your Heart",
             "cholesterol LDL HDL triglycerides atherosclerosis plaque cardiac Malaysia",
             "how LDL deposits in arteries, what raises and lowers LDL, "
             "which Malaysian foods are worst (santan, ghee, offal) and best (fish, oats)"),
            (3, "cardiac_medications",
             "Week {w} — Your Heart Medications Explained",
             "statin aspirin beta blocker ACE inhibitor ARNI cardiac medication food interaction",
             "plain-language explanation of common cardiac medications, why stopping them is dangerous, "
             "and key food interactions (grapefruit with statins, vitamin K with warfarin)"),
            (4, "warning_signs",
             "Week {w} — Warning Signs and When to Call for Help",
             "chest pain angina heart failure warning signs emergency worsening symptoms",
             "recognising red-flag symptoms (new chest pain, sudden breathlessness, leg swelling worsening) "
             "and the exact action to take (ambulance vs. clinic)"),
        ],
        "A": [
            (1, "pulse_log",
             "Week {w} — Daily Pulse and Symptom Log",
             "resting heart rate pulse symptom log cardiac daily monitoring breathlessness fatigue",
             "daily morning resting pulse + symptom check (breathlessness, ankle swelling, fatigue rating)"),
            (2, "activity_tolerance",
             "Week {w} — Activity Tolerance Journal",
             "activity tolerance cardiac heart failure functional capacity daily living log",
             "logging which daily activities cause symptoms, to document improving tolerance over weeks"),
            (3, "medication_adherence",
             "Week {w} — Medication Adherence Tracker",
             "medication adherence cardiac statin antiplatelet compliance tracking habit",
             "simple daily medication tick-off chart, what to do for missed doses, "
             "and why every dose matters for cardiac outcomes"),
            (4, "symptom_review",
             "Week {w} — 4-Week Cardiac Symptom Review",
             "cardiac symptom trend review heart failure NYHA functional class weight",
             "4-week symptom review: weight trend, activity tolerance progression, and when to escalate to doctor"),
        ],
    },
    "PCOS": {
        "condition_tags": ["Polycystic Ovary Syndrome (PCOS)", "Insulin Resistance"],
        "E": [
            (1, "cardio_pcos",
             "Week {w} — Cardio Foundation for PCOS",
             "aerobic cardio exercise PCOS insulin resistance hormones benefit 150 minutes",
             "a 25-min moderate cardio programme for PCOS — how it improves insulin sensitivity, "
             "reduces androgen levels, and supports weight management"),
            (2, "resistance_pcos",
             "Week {w} — Resistance Training for PCOS",
             "resistance strength training PCOS insulin sensitivity lean muscle androgens",
             "why building muscle is especially beneficial for PCOS: improves insulin sensitivity and "
             "reduces testosterone-driven symptoms, with a beginner resistance programme"),
            (3, "light_hiit",
             "Week {w} — Light HIIT for Insulin Sensitivity",
             "HIIT interval training PCOS insulin sensitivity blood sugar exercise benefit",
             "a beginner 20-min light HIIT (30-sec effort / 90-sec rest × 8 rounds) "
             "to maximise insulin-sensitising effects without overtraining cortisol spike in PCOS"),
            (4, "yoga_pcos",
             "Week {w} — Yoga and Stress Reduction for PCOS",
             "yoga stress cortisol PCOS hormonal balance relaxation parasympathetic",
             "yoga sequences to lower cortisol and improve hormonal balance in PCOS: "
             "restorative poses, breathwork, and pelvic-focused stretches"),
        ],
        "K": [
            (1, "understanding_pcos",
             "Week {w} — Understanding PCOS",
             "PCOS polycystic ovary syndrome pathophysiology hormones Malaysia prevalence diagnosis",
             "what PCOS is, how insulin resistance drives hyperandrogenism, diagnosis criteria, "
             "and why it is so common in Malaysian women"),
            (2, "insulin_hormones",
             "Week {w} — Insulin, Hormones, and PCOS",
             "insulin resistance androgens LH FSH ratio PCOS hormones cycle disruption",
             "the LH/FSH ratio disruption in PCOS, how high insulin triggers ovarian androgen production, "
             "and why diet is a first-line treatment"),
            (3, "diet_pcos",
             "Week {w} — The Best Diet for PCOS",
             "low GI diet anti-inflammatory PCOS weight loss insulin sensitivity evidence",
             "evidence review: low-GI diet vs Mediterranean vs low-carb for PCOS, "
             "with Malaysian-friendly meal swaps"),
            (4, "stress_sleep_pcos",
             "Week {w} — Stress, Sleep, and PCOS",
             "cortisol stress sleep quality PCOS hormonal worsening insulin resistance",
             "how chronic stress and poor sleep elevate cortisol → worsen insulin resistance → worsen PCOS: "
             "practical sleep hygiene and stress management"),
        ],
        "A": [
            (1, "cycle_tracking",
             "Week {w} — Menstrual Cycle Tracking",
             "menstrual cycle tracking PCOS irregular periods app symptom log",
             "setting up a period and symptom tracking habit (app or paper): cycle length, spotting, "
             "acne, mood to detect improvement over time"),
            (2, "food_mood_diary",
             "Week {w} — Food and Mood Diary",
             "food diary mood energy PCOS insulin hormones tracking pattern recognition",
             "tracking food choices alongside mood, energy, and cravings to identify insulin-driven patterns"),
            (3, "exercise_energy_log",
             "Week {w} — Exercise and Energy Tracker",
             "exercise energy cravings blood sugar PCOS improvement self-monitoring",
             "logging pre- and post-exercise energy + hunger/cravings to see insulin-sensitising effect"),
            (4, "monthly_symptom_review",
             "Week {w} — Monthly PCOS Symptom Review",
             "PCOS symptom review acne hirsutism fatigue mood cycle weight improvement",
             "monthly review of PCOS symptom domains: cycle regularity, skin, energy, weight, mood — noting trends"),
        ],
    },
    "Dyslipidaemia": {
        "condition_tags": ["Dyslipidaemia", "Hypercholesterolaemia"],
        "E": [
            (1, "aerobic_lipids",
             "Week {w} — Aerobic Exercise and Your Lipid Profile",
             "aerobic exercise LDL HDL triglycerides dyslipidaemia cholesterol reduction mechanism",
             "how aerobic exercise lowers triglycerides and raises HDL: a 25-min walking/cycling programme "
             "with target heart rate zone for lipid benefit"),
            (2, "step_programme",
             "Week {w} — Progressive Daily Step Programme",
             "daily walking steps HDL cholesterol triglyceride improvement goal 7000",
             "progressive daily walking building to 7000-8000 steps/day to improve lipid profile, "
             "with Malaysian-accessible tracking options"),
            (3, "resistance_lipids",
             "Week {w} — Resistance Training for Cholesterol",
             "resistance strength training muscle mass LDL HDL triglycerides metabolic",
             "how resistance training reduces LDL and visceral fat, improves metabolic rate: "
             "a beginner 2×/week programme"),
            (4, "cardio_endurance",
             "Week {w} — Building Cardio Endurance",
             "cardiorespiratory fitness VO2max cholesterol cardiovascular risk dyslipidaemia",
             "4-week cardio progression overview: how VO2max improvement reduces LDL, raises HDL, "
             "and cuts cardiovascular event risk"),
        ],
        "K": [
            (1, "understanding_lipids",
             "Week {w} — Understanding Your Lipid Panel",
             "LDL HDL triglycerides total cholesterol lipid panel interpretation targets Malaysia",
             "what each number means (LDL < 2.6, HDL > 1.0, TG < 1.7), "
             "why total cholesterol alone is misleading, and what optimal looks like"),
            (2, "diet_cholesterol",
             "Week {w} — Diet and Cholesterol",
             "saturated fat trans fat dietary cholesterol LDL raising lowering foods Malaysia",
             "which Malaysian foods raise LDL (santan, ghee, palm oil, processed meat) "
             "and which lower it (oats, psyllium, omega-3 fish, legumes)"),
            (3, "statin_lifestyle",
             "Week {w} — Statins and Lifestyle Together",
             "statin atorvastatin rosuvastatin diet exercise lifestyle dyslipidaemia combination",
             "why lifestyle changes still matter on statins, how they work synergistically, "
             "statin side effects, and grapefruit/alcohol interactions"),
            (4, "cv_risk",
             "Week {w} — Cholesterol and Cardiovascular Risk",
             "LDL cholesterol cardiovascular risk atherosclerosis plaque rupture Malaysia",
             "how LDL builds plaque, what causes plaque rupture (heart attack), "
             "and how reducing LDL by 1 mmol/L cuts risk"),
        ],
        "A": [
            (1, "food_label_audit",
             "Week {w} — Saturated Fat Food Label Audit",
             "food label reading saturated fat trans fat cholesterol Malaysia packaged food",
             "auditing 5 daily-use packaged products for saturated fat content — "
             "calculating how much comes from each"),
            (2, "cooking_tracker",
             "Week {w} — Cooking Method Tracker",
             "cooking method healthy cholesterol steaming grilling baking vs frying switch",
             "tracking this week's cooking methods: how many fried vs. healthier methods, "
             "and identifying one easy swap"),
            (3, "step_goal",
             "Week {w} — Weekly Step Goal Challenge",
             "daily steps aerobic physical activity HDL benefit walking challenge",
             "7-day step challenge: daily goal + check-in, HDL-raising effect of consistent steps"),
            (4, "lipid_trend",
             "Week {w} — Tracking Your Lipid Results Over Time",
             "LDL HDL triglycerides monitoring trend dyslipidaemia progress statin lifestyle",
             "building a simple lipid result log to see trends, understand what changes mean, "
             "and what to tell your doctor"),
        ],
    },
    "General": {
        "condition_tags": [],
        "E": [
            (1, "beginner_fitness",
             "Week {w} — Beginner Fitness Foundation",
             "beginner exercise aerobic walking 150 minutes weekly general wellness",
             "a 3-day beginner programme: 25-min walks + basic stretching, "
             "progressing from sedentary to WHO-recommended 150 min/week"),
            (2, "cardio_build",
             "Week {w} — Building Cardiovascular Fitness",
             "aerobic cardio fitness VO2max cardiorespiratory endurance general wellness",
             "week 2 progression: brisk walking, cycling, or swimming at moderate intensity "
             "to build aerobic base and metabolic health"),
            (3, "strength_basics",
             "Week {w} — Bodyweight Strength Basics",
             "bodyweight exercise strength training squats push-ups lunges general wellness",
             "a 2×/week beginner bodyweight circuit (squat, hinge, push, pull, core) "
             "no equipment, home-friendly for Malaysian adults"),
            (4, "active_recovery",
             "Week {w} — Flexibility and Active Recovery",
             "stretching flexibility yoga active recovery mobility general wellness",
             "full-body stretching + light yoga for active recovery: "
             "why rest days matter, mobility for long-term joint health"),
        ],
        "K": [
            (1, "nutrition_basics",
             "Week {w} — Nutrition Fundamentals",
             "balanced diet macronutrients protein carbohydrate fat micronutrients Malaysia",
             "the basics of balanced nutrition: macros, fibre, micronutrients, "
             "and how Malaysian food can meet or miss the mark"),
            (2, "exercise_science",
             "Week {w} — Why Regular Exercise Is Essential",
             "exercise benefits physical activity heart brain metabolic mental health longevity Malaysia",
             "the evidence: 150 min/week moderate exercise reduces all-cause mortality, "
             "mechanism across cardiovascular, metabolic, and mental health"),
            (3, "sleep_health",
             "Week {w} — Sleep and Your Health",
             "sleep quality duration health metabolism weight cortisol immune function",
             "how poor sleep drives weight gain, impairs insulin sensitivity, weakens immunity, "
             "and harms mental health — practical sleep hygiene for Malaysians"),
            (4, "stress_disease",
             "Week {w} — Chronic Stress and Disease",
             "chronic stress cortisol inflammation metabolic disease cardiovascular risk lifestyle",
             "how long-term stress elevates cortisol → chronic inflammation → increased risk of diabetes, "
             "heart disease, and depression; evidence-based interventions"),
        ],
        "A": [
            (1, "step_challenge",
             "Week {w} — 7-Day Step Challenge",
             "daily steps 8000 10000 walking activity wellness habit tracking",
             "a 7-day progressive step challenge: start at personal baseline, add 500 steps/day, "
             "aiming toward 8000-10000 by day 7"),
            (2, "hydration_habit",
             "Week {w} — Daily Hydration Habit",
             "water intake hydration 2L daily tracking health Malaysia climate heat",
             "tracking daily water intake toward 2-2.5L/day in Malaysian heat: "
             "visual tracker, replacing sugary drinks, recognising dehydration signs"),
            (3, "sleep_log",
             "Week {w} — Sleep Quality Log",
             "sleep tracking bedtime wake time quality rating wellness habit",
             "7-day sleep log: bedtime, wake time, quality rating 1-5, and one habit to improve it"),
            (4, "wellness_checkin",
             "Week {w} — 4-Week Wellness Check-In",
             "wellness review energy mood weight activity sleep nutrition progress goals",
             "structured 4-week review across 5 domains: steps, sleep, nutrition, stress, and weight — "
             "setting realistic next-month goals"),
        ],
    },
}


# ---------------------------------------------------------------------------
# Build flat niche case list from templates
# ---------------------------------------------------------------------------

def build_eka_cases(iso_week: int) -> list:
    """Return 84 niche case dicts for the given ISO week number."""
    rotation = ((iso_week - 1) % 4) + 1  # 1-4 rotation
    cases = []
    for group, gdata in _TEMPLATES.items():
        for ctype in ("E", "K", "A"):
            entries = gdata[ctype]
            slot, topic_slug, title_tmpl, rag_query, prompt_topic = entries[rotation - 1]
            cases.append({
                "group":          group,
                "condition_tags": gdata["condition_tags"],
                "content_type":   ctype,
                "week_number":    iso_week,
                "topic":          topic_slug,
                "title":          title_tmpl.format(w=iso_week),
                "rag_query":      rag_query,
                "prompt_topic":   prompt_topic,
            })
    return cases


# ---------------------------------------------------------------------------
# RAG retrieval helper (shared with generate_content.py)
# ---------------------------------------------------------------------------

def _retrieve_chunks(query: str, client_id: int, top_k: int = 8) -> str:
    from langchain_community.vectorstores import PGVector
    from vector_store import get_connection_string
    from embeddings import get_embedding_function

    conn = get_connection_string()
    emb = get_embedding_function()
    base_db = PGVector(connection_string=conn, embedding_function=emb,
                       collection_name="base_knowledge", use_jsonb=True)
    docs = base_db.similarity_search(query, k=top_k)
    try:
        client_db = PGVector(connection_string=conn, embedding_function=emb,
                             collection_name=f"client_{client_id}_knowledge", use_jsonb=True)
        seen = {d.page_content for d in docs}
        for d in client_db.similarity_search(query, k=3):
            if d.page_content not in seen:
                docs.append(d)
    except Exception:
        pass
    return "\n\n---\n\n".join(d.page_content[:800] for d in docs[:top_k])


# ---------------------------------------------------------------------------
# Type-specific LLM generation
# ---------------------------------------------------------------------------

_PERSONALIZATION_GUIDANCE = """\
Personalization level rules:
L0 (no risk, general wellness): Full spectrum; vigorous activity allowed; no clinical stop signs needed.
L1 (emerging/moderate risk): Structured, safety-aware; clear do/don't boundaries; moderate intensity max.
L2 (established conditions, higher CV risk): Low-intensity only; symptom monitoring required; strict stop conditions.
L3 (high clinical risk, recent cardiac event): Medical oversight only; extremely gentle; include emergency guidance.
"""


def _generate_exercise(niche: dict, chunks: str) -> dict:
    from llm import call_ollama_generate
    prompt = f"""You are a clinical exercise physiologist creating cardiac and metabolic rehabilitation content for a Malaysian hospital.

TOPIC: {niche["prompt_topic"]}
CONDITION GROUP: {niche["group"]}
WEEK: {niche["week_number"]}

{_PERSONALIZATION_GUIDANCE}

CLINICAL EVIDENCE:
{chunks or "No specific guideline chunks available — use evidence-based clinical knowledge."}

TASK: Generate a structured exercise session plan. Return ONLY valid JSON — no prose, no markdown fences:
{{
  "exercise_type": "aerobic | resistance | flexibility | balance | combined",
  "duration_min": <integer>,
  "frequency_per_week": <integer>,
  "intensity": "very light | light | moderate | vigorous",
  "warmup": ["step 1", "step 2", "step 3"],
  "main_activity": ["step 1", "step 2", "step 3", "step 4"],
  "cooldown": ["step 1", "step 2"],
  "level_modifications": {{
    "L0": "modification for general/well patients",
    "L1": "modification for emerging risk",
    "L2": "modification for established conditions",
    "L3": "modification for high-risk/post-cardiac-event patients"
  }},
  "safety_stop_signs": ["sign 1", "sign 2", "sign 3"],
  "equipment_needed": ["item 1", "item 2"],
  "malaysian_context": "any Malaysia-specific notes (heat, shoes, local venues, etc.)"
}}"""
    return _call_and_parse(prompt, 1000)


def _generate_knowledge(niche: dict, chunks: str) -> dict:
    from llm import call_ollama_generate
    prompt = f"""You are a clinical educator creating patient health literacy content for a Malaysian hospital.

TOPIC: {niche["prompt_topic"]}
CONDITION GROUP: {niche["group"]}
WEEK: {niche["week_number"]}

CLINICAL EVIDENCE:
{chunks or "No specific guideline chunks available — use evidence-based clinical knowledge."}

TASK: Generate 6 educational learning points for patients. Each point should be clear, jargon-free, actionable, and culturally relevant to Malaysia. Return ONLY valid JSON — no prose, no markdown fences:
{{
  "topic_summary": "one sentence summarising the topic",
  "learning_points": [
    {{"point": "short heading", "explanation": "1-2 sentence explanation", "why_it_matters": "why this matters to patient"}},
    {{"point": "...", "explanation": "...", "why_it_matters": "..."}},
    {{"point": "...", "explanation": "...", "why_it_matters": "..."}},
    {{"point": "...", "explanation": "...", "why_it_matters": "..."}},
    {{"point": "...", "explanation": "...", "why_it_matters": "..."}},
    {{"point": "...", "explanation": "...", "why_it_matters": "..."}}
  ],
  "key_takeaway": "one-sentence bottom line for the patient",
  "local_context": "specific Malaysian food, culture, or healthcare context note"
}}"""
    return _call_and_parse(prompt, 900)


def _generate_activity(niche: dict, chunks: str) -> dict:
    from llm import call_ollama_generate
    prompt = f"""You are a health behaviour coach creating patient habit-building tasks for a Malaysian hospital programme.

TOPIC: {niche["prompt_topic"]}
CONDITION GROUP: {niche["group"]}
WEEK: {niche["week_number"]}

TASK: Design a practical weekly behavioural activity. It must be simple enough to do daily, relevant to Malaysian patients, and directly support health outcomes. Return ONLY valid JSON — no prose, no markdown fences:
{{
  "task_name": "short catchy name",
  "description": "1-2 sentence description of the task",
  "instructions": ["step 1", "step 2", "step 3", "step 4"],
  "tracking_method": "how to track (app / paper chart / phone notes)",
  "weekly_goal": "specific measurable goal for this week",
  "micro_actions": [
    "Monday: ...",
    "Tuesday: ...",
    "Wednesday: ...",
    "Thursday: ...",
    "Friday: ...",
    "Weekend: ..."
  ],
  "self_monitoring_prompts": [
    "End-of-day reflection question 1",
    "End-of-day reflection question 2"
  ],
  "success_looks_like": "what completing this week successfully looks like"
}}"""
    return _call_and_parse(prompt, 800)


def _call_and_parse(prompt: str, max_tokens: int) -> dict:
    from llm import call_ollama_generate
    raw = call_ollama_generate(prompt, max_tokens=max_tokens).strip()
    if raw.startswith("```"):
        start = raw.find("{")
        end = raw.rfind("}") + 1
        raw = raw[start:end] if start != -1 else raw
    try:
        result = json.loads(raw)
        if isinstance(result, dict):
            return result
    except json.JSONDecodeError:
        pass
    return {"raw_output": raw[:1000], "parse_error": True}


_GENERATORS = {"E": _generate_exercise, "K": _generate_knowledge, "A": _generate_activity}


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

def _write_eka_excel(results: list, output_path: str):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    hdr_fill  = PatternFill("solid", fgColor="1a3c6b")
    hdr_font  = Font(bold=True, color="FFFFFF", size=11)
    type_fills = {
        "E": PatternFill("solid", fgColor="d4e6f1"),
        "K": PatternFill("solid", fgColor="d5f5e3"),
        "A": PatternFill("solid", fgColor="fef9e7"),
    }
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )

    type_labels = {"E": "Exercise", "K": "Knowledge", "A": "Activity"}

    for ctype in ("E", "K", "A"):
        ws = wb.create_sheet(title=type_labels[ctype])
        headers = ["Group", "Week", "Topic", "Title", "Field", "Value", "Status"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin
        ws.row_dimensions[1].height = 22

        row = 2
        items = [r for r in results if r["content_type"] == ctype]
        for item in sorted(items, key=lambda x: x["group"]):
            content = item.get("content", {})
            if not content:
                content = {"(empty)": "(generation failed)"}
            fill = type_fills[ctype]
            for field, value in content.items():
                if isinstance(value, (list, dict)):
                    value = json.dumps(value, ensure_ascii=False)
                for col, val in enumerate([
                    item["group"], item["week_number"],
                    item["topic"], item["title"],
                    field, str(value), "raw — pending review"
                ], 1):
                    cell = ws.cell(row=row, column=col, value=val)
                    cell.border = thin
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                    if col <= 4:
                        cell.fill = fill
                row += 1

        for i, w in enumerate([14, 8, 22, 40, 25, 70, 20], 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A2"

    wb.save(output_path)
    print(f"  Excel saved → {output_path}")


# ---------------------------------------------------------------------------
# Main generation function
# ---------------------------------------------------------------------------

def generate_weekly_eka(iso_week: int = None, client_id: int = 4,
                        filter_group: str = None, filter_type: str = None,
                        dry_run: bool = False, force: bool = False,
                        output_dir: str = None) -> list:
    if iso_week is None:
        iso_week = date.today().isocalendar()[1]

    if output_dir is None:
        output_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "materials")
    os.makedirs(output_dir, exist_ok=True)

    cases = build_eka_cases(iso_week)
    if filter_group:
        cases = [c for c in cases if c["group"] == filter_group]
    if filter_type:
        cases = [c for c in cases if c["content_type"] == filter_type]

    if not cases:
        print("No cases match filters.")
        return []

    rotation = ((iso_week - 1) % 4) + 1
    print(f"\n{'[DRY RUN] ' if dry_run else ''}Weekly EKA Generation — ISO Week {iso_week} (Rotation {rotation}/4)")
    print(f"  Generating {len(cases)} items  |  Client: {client_id}  |  Force: {force}\n")

    db_session = None
    if not dry_run:
        import database as db_module
        db_module.create_db_and_tables()
        db_session = db_module.SessionLocal()

    results = []
    try:
        for i, niche in enumerate(cases, 1):
            label = f"[{i}/{len(cases)}] {niche['group']} / {niche['content_type']} / {niche['topic']}"
            print(f"  {label}")

            if dry_run:
                results.append({**niche, "content": {}, "skipped": True})
                print("    → (dry run)")
                continue

            print("    → retrieving chunks...")
            try:
                chunks = _retrieve_chunks(niche["rag_query"], client_id)
            except Exception as e:
                print(f"    → retrieval error: {e}")
                chunks = ""

            print("    → generating via Ollama...")
            try:
                gen_fn = _GENERATORS[niche["content_type"]]
                content = gen_fn(niche, chunks)
            except Exception as e:
                print(f"    → generation error: {e}")
                content = {"error": str(e)}

            parse_ok = not content.get("parse_error")
            print(f"    → {'OK' if parse_ok else 'PARSE ERROR'} ({len(str(content))} chars)")

            result = {**niche, "content": content}
            results.append(result)

            if db_session:
                try:
                    db_module.upsert_eka_material(
                        db_session,
                        condition_group=niche["group"],
                        condition_tags=niche["condition_tags"],
                        content_type=niche["content_type"],
                        week_number=niche["week_number"],
                        topic=niche["topic"],
                        title=niche["title"],
                        raw_content=content,
                        force=force,
                    )
                except Exception as e:
                    print(f"    → DB write error: {e}")

    finally:
        if db_session:
            db_session.close()

    if not dry_run and results:
        ts = datetime.now().strftime("%Y-%m-%d")
        excel_path = os.path.join(output_dir, f"eka_week{iso_week}_{ts}.xlsx")
        try:
            _write_eka_excel(results, excel_path)
        except Exception as e:
            print(f"  Excel write error: {e}")

    print(f"\nDone. {len(results)} items generated for week {iso_week}.")
    return results


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Generate weekly EKA content library")
    parser.add_argument("--week",      type=int, help="ISO week number (default: current week)")
    parser.add_argument("--client-id", type=int, default=4)
    parser.add_argument("--group",     type=str, help="Filter to one group (T2DM, HTN, CKD, Cardiac, PCOS, Dyslipidaemia, General)")
    parser.add_argument("--type",      type=str, help="Filter to one type (E, K, A)")
    parser.add_argument("--dry-run",   action="store_true")
    parser.add_argument("--force",     action="store_true", help="Overwrite existing DB rows")
    parser.add_argument("--output-dir", type=str)
    args = parser.parse_args()

    generate_weekly_eka(
        iso_week=args.week,
        client_id=args.client_id,
        filter_group=args.group,
        filter_type=args.type,
        dry_run=args.dry_run,
        force=args.force,
        output_dir=args.output_dir,
    )
