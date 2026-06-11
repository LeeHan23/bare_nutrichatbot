"""
generate_content.py — Build the educational content library.

For each niche case (condition group × day offset × topic) this script:
  1. Runs a targeted RAG query against the knowledge base
  2. Calls Ollama (qwen2.5:32b) to generate 6 structured tips grounded in evidence
  3. Stores tips in the content_materials DB table (is_active=False until dev team polishes)
  4. Exports everything to an Excel workbook for the dev team

Usage:
    # Generate all niche cases for client 4
    python scripts/generate_content.py --client-id 4

    # Generate only one niche case (fast test)
    python scripts/generate_content.py --client-id 4 --group T2DM --day 3

    # Dry run — print what would be generated without calling LLM or writing DB
    python scripts/generate_content.py --client-id 4 --dry-run

    # Skip DB writes, only produce the Excel
    python scripts/generate_content.py --client-id 4 --no-db
"""
import argparse
import json
import os
import sys
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

# ---------------------------------------------------------------------------
# Niche case definitions
# Each entry = one sheet section in the Excel and one DB row in content_materials
# ---------------------------------------------------------------------------

NICHE_CASES = [
    # ── T2DM ──────────────────────────────────────────────────────────────
    {
        "group": "T2DM", "condition_tags": ["Type 2 Diabetes"],
        "day_offset": 3, "topic": "breakfast_choices",
        "title": "Day 3 — Breakfast Choices for Diabetes",
        "rag_query": "best breakfast foods type 2 diabetes Malaysia low glycaemic index",
        "prompt_topic": "choosing a diabetes-friendly breakfast in Malaysia (nasi lemak, roti, mee goreng alternatives)",
    },
    {
        "group": "T2DM", "condition_tags": ["Type 2 Diabetes"],
        "day_offset": 5, "topic": "local_food_swaps",
        "title": "Day 5 — Malaysian Food Swaps for Diabetes",
        "rag_query": "Malaysian food substitutions diabetes healthier alternatives hawker food",
        "prompt_topic": "swapping common Malaysian hawker and home-cooked foods for better blood sugar control",
    },
    {
        "group": "T2DM", "condition_tags": ["Type 2 Diabetes"],
        "day_offset": 7, "topic": "medication_food_timing",
        "title": "Day 7 — Metformin and Meal Timing",
        "rag_query": "metformin medication food timing diabetes diet interaction side effects",
        "prompt_topic": "how to time meals around Metformin to reduce side effects and stabilise blood sugar",
    },
    {
        "group": "T2DM", "condition_tags": ["Type 2 Diabetes"],
        "day_offset": 14, "topic": "portion_control",
        "title": "Day 14 — Portion Control (Suku-Suku-Separuh)",
        "rag_query": "portion control diabetes plate method carbohydrate serving size Malaysia",
        "prompt_topic": "using the Malaysian suku-suku-separuh plate method to manage carbohydrate intake",
    },
    {
        "group": "T2DM", "condition_tags": ["Type 2 Diabetes"],
        "day_offset": 21, "topic": "eating_out",
        "title": "Day 21 — Eating Out Safely with Diabetes",
        "rag_query": "eating out restaurant hawker stall diabetes safe food choices Malaysia",
        "prompt_topic": "eating safely at Malaysian restaurants, mamak stalls, and hawker centres with diabetes",
    },
    {
        "group": "T2DM", "condition_tags": ["Type 2 Diabetes"],
        "day_offset": 30, "topic": "progress_check",
        "title": "Day 30 — 30-Day Progress Check",
        "rag_query": "diabetes diet monitoring HbA1c blood sugar self-monitoring goals",
        "prompt_topic": "what to track and celebrate at the 30-day mark for diabetes dietary management",
    },

    # ── HTN ───────────────────────────────────────────────────────────────
    {
        "group": "HTN", "condition_tags": ["Hypertension"],
        "day_offset": 3, "topic": "sodium_basics",
        "title": "Day 3 — Understanding Sodium and Blood Pressure",
        "rag_query": "sodium reduction hypertension blood pressure diet DASH Malaysia",
        "prompt_topic": "how sodium raises blood pressure and which common Malaysian foods are highest in sodium",
    },
    {
        "group": "HTN", "condition_tags": ["Hypertension"],
        "day_offset": 5, "topic": "hidden_sodium",
        "title": "Day 5 — Hidden Sodium in Malaysian Foods",
        "rag_query": "hidden salt processed food soy sauce belacan kicap sodium hypertension",
        "prompt_topic": "identifying hidden sodium in Malaysian cooking sauces, condiments, and processed foods",
    },
    {
        "group": "HTN", "condition_tags": ["Hypertension"],
        "day_offset": 7, "topic": "dash_diet",
        "title": "Day 7 — DASH Diet Adapted for Malaysia",
        "rag_query": "DASH diet hypertension potassium magnesium fruits vegetables Malaysia",
        "prompt_topic": "adapting the DASH diet to Malaysian food culture — what to increase, what to reduce",
    },
    {
        "group": "HTN", "condition_tags": ["Hypertension"],
        "day_offset": 14, "topic": "beverages",
        "title": "Day 14 — Beverages and Blood Pressure",
        "rag_query": "alcohol caffeine teh tarik kopi beverages blood pressure hypertension",
        "prompt_topic": "how Teh Tarik, kopi, energy drinks, and alcohol affect blood pressure",
    },
    {
        "group": "HTN", "condition_tags": ["Hypertension"],
        "day_offset": 21, "topic": "eating_out",
        "title": "Day 21 — Low-Sodium Eating Out Guide",
        "rag_query": "low sodium eating out restaurant hypertension food choices Malaysia",
        "prompt_topic": "ordering lower-sodium meals at Malaysian restaurants, kopitiam, and food courts",
    },
    {
        "group": "HTN", "condition_tags": ["Hypertension"],
        "day_offset": 30, "topic": "progress_check",
        "title": "Day 30 — Blood Pressure Monitoring and Diet Progress",
        "rag_query": "blood pressure self-monitoring diet progress hypertension goals",
        "prompt_topic": "what dietary changes to track and how to interpret home blood pressure readings",
    },

    # ── CKD ───────────────────────────────────────────────────────────────
    {
        "group": "CKD", "condition_tags": ["Chronic Kidney Disease"],
        "day_offset": 3, "topic": "fluid_restriction",
        "title": "Day 3 — Managing Fluid Intake with CKD",
        "rag_query": "fluid restriction chronic kidney disease CKD daily intake management",
        "prompt_topic": "how to measure and limit fluid intake practically for CKD patients in Malaysia",
    },
    {
        "group": "CKD", "condition_tags": ["Chronic Kidney Disease"],
        "day_offset": 5, "topic": "potassium_foods",
        "title": "Day 5 — High-Potassium Foods to Avoid",
        "rag_query": "high potassium foods CKD kidney disease avoid banana coconut water Malaysia",
        "prompt_topic": "which Malaysian foods are high in potassium and why CKD patients must limit them",
    },
    {
        "group": "CKD", "condition_tags": ["Chronic Kidney Disease"],
        "day_offset": 7, "topic": "phosphorus",
        "title": "Day 7 — Phosphorus and Kidney Health",
        "rag_query": "phosphorus restriction CKD kidney disease dairy nuts seeds phosphate binders",
        "prompt_topic": "understanding phosphorus in the diet and which Malaysian foods to reduce or avoid",
    },
    {
        "group": "CKD", "condition_tags": ["Chronic Kidney Disease"],
        "day_offset": 14, "topic": "protein_balance",
        "title": "Day 14 — How Much Protein is Right for You?",
        "rag_query": "protein intake CKD kidney disease restriction adequate nutrition dialysis",
        "prompt_topic": "protein balance for CKD — enough to maintain muscle but not overload the kidneys",
    },
    {
        "group": "CKD", "condition_tags": ["Chronic Kidney Disease"],
        "day_offset": 21, "topic": "eating_out",
        "title": "Day 21 — Eating Out Safely with CKD",
        "rag_query": "CKD kidney disease eating out restaurant food choices safe low potassium sodium",
        "prompt_topic": "navigating Malaysian restaurants and hawker food with CKD dietary restrictions",
    },
    {
        "group": "CKD", "condition_tags": ["Chronic Kidney Disease"],
        "day_offset": 30, "topic": "progress_check",
        "title": "Day 30 — Kidney Diet Review",
        "rag_query": "CKD diet monitoring eGFR lab values potassium phosphorus sodium progress",
        "prompt_topic": "which lab values to watch and how diet connects to kidney function test results",
    },

    # ── Cardiac ───────────────────────────────────────────────────────────
    {
        "group": "Cardiac", "condition_tags": ["Ischaemic Heart Disease", "Heart Failure", "Dyslipidaemia"],
        "day_offset": 3, "topic": "fat_types",
        "title": "Day 3 — Good Fats vs. Bad Fats for Your Heart",
        "rag_query": "saturated fat unsaturated fat omega-3 heart disease cholesterol Malaysia",
        "prompt_topic": "the difference between saturated fats (coconut milk, ghee, palm oil) and heart-healthy unsaturated fats",
    },
    {
        "group": "Cardiac", "condition_tags": ["Ischaemic Heart Disease", "Heart Failure", "Dyslipidaemia"],
        "day_offset": 5, "topic": "local_food_swaps",
        "title": "Day 5 — Heart-Healthy Malaysian Food Swaps",
        "rag_query": "heart healthy Malaysian food alternatives cooking methods santan coconut milk",
        "prompt_topic": "swapping santan dishes, fried foods, and high-cholesterol Malaysian foods for heart-healthier options",
    },
    {
        "group": "Cardiac", "condition_tags": ["Ischaemic Heart Disease", "Heart Failure", "Dyslipidaemia"],
        "day_offset": 7, "topic": "cardiac_medications_diet",
        "title": "Day 7 — Your Heart Medications and What to Eat",
        "rag_query": "statin medication diet interaction grapefruit cholesterol heart disease warfarin",
        "prompt_topic": "how cardiac medications (statins, aspirin, beta-blockers) interact with food",
    },
    {
        "group": "Cardiac", "condition_tags": ["Ischaemic Heart Disease", "Heart Failure", "Dyslipidaemia"],
        "day_offset": 14, "topic": "sodium_cardiac",
        "title": "Day 14 — Sodium and Your Heart",
        "rag_query": "low sodium cardiac diet heart failure fluid retention blood pressure",
        "prompt_topic": "why sodium restriction is especially critical for cardiac patients — fluid, swelling, and BP",
    },
    {
        "group": "Cardiac", "condition_tags": ["Ischaemic Heart Disease", "Heart Failure", "Dyslipidaemia"],
        "day_offset": 21, "topic": "eating_out",
        "title": "Day 21 — Eating Out with a Heart Condition",
        "rag_query": "heart healthy eating out restaurant cardiac diet low fat sodium Malaysia",
        "prompt_topic": "making heart-safe choices at Malaysian restaurants, weddings, and family gatherings",
    },
    {
        "group": "Cardiac", "condition_tags": ["Ischaemic Heart Disease", "Heart Failure", "Dyslipidaemia"],
        "day_offset": 30, "topic": "progress_check",
        "title": "Day 30 — Cardiac Diet Progress Review",
        "rag_query": "cardiac diet progress cholesterol LDL weight blood pressure monitoring goals",
        "prompt_topic": "which numbers to track at the 30-day mark and what cardiac dietary success looks like",
    },

    # ── PCOS ──────────────────────────────────────────────────────────────
    {
        "group": "PCOS", "condition_tags": ["Polycystic Ovary Syndrome (PCOS)", "Insulin Resistance"],
        "day_offset": 3, "topic": "low_gi_basics",
        "title": "Day 3 — Low-GI Eating for PCOS",
        "rag_query": "low glycaemic index diet PCOS insulin resistance blood sugar management",
        "prompt_topic": "why low-GI foods help PCOS and which Malaysian foods have a high vs. low glycaemic index",
    },
    {
        "group": "PCOS", "condition_tags": ["Polycystic Ovary Syndrome (PCOS)", "Insulin Resistance"],
        "day_offset": 5, "topic": "anti_inflammatory",
        "title": "Day 5 — Anti-Inflammatory Foods for PCOS",
        "rag_query": "anti-inflammatory diet PCOS omega-3 berries vegetables turmeric haldi",
        "prompt_topic": "incorporating anti-inflammatory foods available in Malaysia to help with PCOS symptoms",
    },
    {
        "group": "PCOS", "condition_tags": ["Polycystic Ovary Syndrome (PCOS)", "Insulin Resistance"],
        "day_offset": 7, "topic": "meal_timing",
        "title": "Day 7 — Meal Timing and PCOS",
        "rag_query": "meal timing breakfast skipping irregular meals PCOS insulin cortisol",
        "prompt_topic": "why skipping breakfast worsens PCOS symptoms and how to structure meals for better hormone balance",
    },
    {
        "group": "PCOS", "condition_tags": ["Polycystic Ovary Syndrome (PCOS)", "Insulin Resistance"],
        "day_offset": 14, "topic": "weight_management",
        "title": "Day 14 — Weight and PCOS",
        "rag_query": "weight management PCOS BMI insulin sensitivity diet exercise",
        "prompt_topic": "how even modest weight loss improves PCOS symptoms and which dietary changes are most effective",
    },
    {
        "group": "PCOS", "condition_tags": ["Polycystic Ovary Syndrome (PCOS)", "Insulin Resistance"],
        "day_offset": 21, "topic": "eating_out",
        "title": "Day 21 — Eating Out with PCOS",
        "rag_query": "low GI eating out restaurant PCOS insulin resistance Malaysia food choices",
        "prompt_topic": "low-GI choices at Malaysian food courts, restaurants, and social events with PCOS",
    },
    {
        "group": "PCOS", "condition_tags": ["Polycystic Ovary Syndrome (PCOS)", "Insulin Resistance"],
        "day_offset": 30, "topic": "progress_check",
        "title": "Day 30 — PCOS Nutrition Check-In",
        "rag_query": "PCOS diet progress symptoms tracking insulin weight hormones",
        "prompt_topic": "what to observe at 30 days — symptoms, weight, energy, and mood as PCOS diet indicators",
    },

    # ── Dyslipidaemia ─────────────────────────────────────────────────────
    {
        "group": "Dyslipidaemia", "condition_tags": ["Dyslipidaemia", "Hypercholesterolaemia"],
        "day_offset": 3, "topic": "cholesterol_foods",
        "title": "Day 3 — Foods That Affect Your Cholesterol",
        "rag_query": "cholesterol diet saturated fat trans fat LDL HDL dyslipidaemia food",
        "prompt_topic": "which Malaysian foods raise LDL cholesterol and which help lower it",
    },
    {
        "group": "Dyslipidaemia", "condition_tags": ["Dyslipidaemia", "Hypercholesterolaemia"],
        "day_offset": 5, "topic": "fibre_sterols",
        "title": "Day 5 — Fibre and Plant Sterols",
        "rag_query": "soluble fibre oats plant sterols LDL cholesterol reduction dyslipidaemia",
        "prompt_topic": "how oats, legumes, and fibre-rich foods lower LDL and where to find them in Malaysia",
    },
    {
        "group": "Dyslipidaemia", "condition_tags": ["Dyslipidaemia", "Hypercholesterolaemia"],
        "day_offset": 7, "topic": "statin_diet",
        "title": "Day 7 — Statins and Your Diet",
        "rag_query": "statin atorvastatin rosuvastatin diet interaction grapefruit alcohol cholesterol",
        "prompt_topic": "what to eat and avoid while on statins, including grapefruit and alcohol interactions",
    },
    {
        "group": "Dyslipidaemia", "condition_tags": ["Dyslipidaemia", "Hypercholesterolaemia"],
        "day_offset": 14, "topic": "cooking_methods",
        "title": "Day 14 — Healthier Cooking Methods",
        "rag_query": "healthy cooking methods grilling steaming air fry cholesterol saturated fat reduction",
        "prompt_topic": "replacing deep frying with steaming, air frying, and grilling in Malaysian home cooking",
    },
    {
        "group": "Dyslipidaemia", "condition_tags": ["Dyslipidaemia", "Hypercholesterolaemia"],
        "day_offset": 21, "topic": "eating_out",
        "title": "Day 21 — Eating Out with High Cholesterol",
        "rag_query": "low fat low cholesterol eating out Malaysia restaurant hawker food",
        "prompt_topic": "lower-cholesterol choices at Malaysian hawker stalls, mamak restaurants, and Chinese kopitiam",
    },
    {
        "group": "Dyslipidaemia", "condition_tags": ["Dyslipidaemia", "Hypercholesterolaemia"],
        "day_offset": 30, "topic": "progress_check",
        "title": "Day 30 — Cholesterol Diet Progress",
        "rag_query": "LDL cholesterol progress diet monitoring triglycerides HDL goals",
        "prompt_topic": "what a 30-day cholesterol-lowering diet can realistically achieve and how to interpret lipid panel results",
    },

    # ── General Wellness (L0) ─────────────────────────────────────────────
    {
        "group": "General", "condition_tags": [],
        "day_offset": 3, "topic": "balanced_plate",
        "title": "Day 3 — Building a Balanced Malaysian Plate",
        "rag_query": "balanced diet Malaysia suku suku separuh healthy eating guidelines",
        "prompt_topic": "building a balanced plate using Malaysian foods and the suku-suku-separuh method",
    },
    {
        "group": "General", "condition_tags": [],
        "day_offset": 5, "topic": "hydration",
        "title": "Day 5 — Staying Well-Hydrated",
        "rag_query": "water intake hydration health Malaysia daily fluid recommendation",
        "prompt_topic": "daily water and fluid intake in the Malaysian climate — how much, what counts, what to reduce",
    },
    {
        "group": "General", "condition_tags": [],
        "day_offset": 7, "topic": "nutrition_labels",
        "title": "Day 7 — Reading Food Labels",
        "rag_query": "food label reading nutrition facts serving size sodium sugar Malaysia",
        "prompt_topic": "how to read Malaysian food labels — serving size, sodium, sugar, and fat content",
    },
    {
        "group": "General", "condition_tags": [],
        "day_offset": 14, "topic": "performance_nutrition",
        "title": "Day 14 — Eating for Energy and Exercise",
        "rag_query": "sports nutrition pre workout post workout meal energy carbohydrate protein",
        "prompt_topic": "fuelling exercise and active lifestyle with Malaysian food — pre and post-workout nutrition",
    },
    {
        "group": "General", "condition_tags": [],
        "day_offset": 21, "topic": "eating_out",
        "title": "Day 21 — Healthy Eating Out in Malaysia",
        "rag_query": "healthy eating out Malaysia hawker food balanced choices restaurant",
        "prompt_topic": "making healthier choices when eating out at Malaysian food courts and restaurants",
    },
    {
        "group": "General", "condition_tags": [],
        "day_offset": 30, "topic": "progress_check",
        "title": "Day 30 — Wellness Check-In",
        "rag_query": "general wellness nutrition progress energy weight wellbeing goals",
        "prompt_topic": "reviewing nutrition habits at 30 days and setting realistic next-month goals",
    },
]

# ---------------------------------------------------------------------------
# Condition group → patient condition keywords (for matching in scheduler)
# ---------------------------------------------------------------------------

CONDITION_MAP = {
    "T2DM":         ["diabetes", "t2dm", "type 2"],
    "HTN":          ["hypertension"],
    "CKD":          ["kidney disease", "ckd", "renal"],
    "Cardiac":      ["heart disease", "ischaemic", "ihd", "cabg", "heart failure",
                     "dyslipidaemia", "hypercholesterol"],
    "PCOS":         ["pcos", "polycystic ovary", "insulin resistance"],
    "Dyslipidaemia":["dyslipidaemia", "hypercholesterol", "lipid"],
    "General":      [],  # fallback — matches patients with no conditions
}

SCHEDULE_DAYS = [3, 5, 7, 14, 21, 30]


def conditions_to_groups(conditions: list) -> list:
    """Map a patient's condition strings to niche case group names."""
    if not conditions:
        return ["General"]
    groups = set()
    lowered = [c.lower() for c in conditions]
    for group, keywords in CONDITION_MAP.items():
        if not keywords:
            continue
        for kw in keywords:
            if any(kw in c for c in lowered):
                groups.add(group)
                break
    return list(groups) if groups else ["General"]


# ---------------------------------------------------------------------------
# RAG retrieval helper
# ---------------------------------------------------------------------------

def _retrieve_chunks(query: str, client_id: int, top_k: int = 8) -> str:
    from langchain_community.vectorstores import PGVector
    from vector_store import get_connection_string
    from embeddings import get_embedding_function

    conn = get_connection_string()
    emb = get_embedding_function()

    base_db = PGVector(
        connection_string=conn,
        embedding_function=emb,
        collection_name="base_knowledge",
        use_jsonb=True,
    )
    docs = base_db.similarity_search(query, k=top_k)

    # Also pull from client knowledge if available
    try:
        client_db = PGVector(
            connection_string=conn,
            embedding_function=emb,
            collection_name=f"client_{client_id}_knowledge",
            use_jsonb=True,
        )
        seen = {d.page_content for d in docs}
        for d in client_db.similarity_search(query, k=3):
            if d.page_content not in seen:
                docs.append(d)
    except Exception:
        pass

    return "\n\n---\n\n".join(d.page_content[:800] for d in docs[:top_k])


# ---------------------------------------------------------------------------
# Tip generation
# ---------------------------------------------------------------------------

def _generate_tips(niche: dict, chunks: str) -> list:
    """Call Ollama to generate 6 structured tips. Returns list of dicts."""
    from llm import call_ollama_generate

    prompt = f"""You are a clinical dietitian creating patient education materials for a Malaysian hospital.

TOPIC: {niche["prompt_topic"]}
CONDITION GROUP: {niche["group"]}
DAY IN PROGRAMME: Day {niche["day_offset"]}

CLINICAL EVIDENCE (from guidelines):
{chunks if chunks else "No specific guideline chunks available — use general clinical knowledge."}

TASK: Write exactly 6 practical, actionable tips for this topic.
Each tip must:
- Be 1-3 sentences
- Reference Malaysian food culture where relevant (local dishes, ingredients, brand names)
- Be specific and immediately actionable
- Be appropriate for a patient at Day {niche["day_offset"]} of their nutrition programme

Return ONLY a valid JSON array. No prose before or after. No markdown fences:
[
  {{"tip_number": 1, "tip": "...", "source_hint": "clinical guideline or general knowledge"}},
  {{"tip_number": 2, "tip": "...", "source_hint": "..."}},
  {{"tip_number": 3, "tip": "...", "source_hint": "..."}},
  {{"tip_number": 4, "tip": "...", "source_hint": "..."}},
  {{"tip_number": 5, "tip": "...", "source_hint": "..."}},
  {{"tip_number": 6, "tip": "...", "source_hint": "..."}}
]"""

    raw = call_ollama_generate(prompt, max_tokens=800)

    # Strip markdown fences if present
    raw = raw.strip()
    if raw.startswith("```"):
        raw = raw[raw.find("["):]
    if raw.endswith("```"):
        raw = raw[:raw.rfind("]") + 1]

    try:
        tips = json.loads(raw)
        if isinstance(tips, list) and len(tips) > 0:
            return tips
    except json.JSONDecodeError:
        pass

    # Fallback: wrap raw text as a single tip
    return [{"tip_number": 1, "tip": raw[:500], "source_hint": "generated"}]


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

def _write_excel(all_results: list, output_path: str):
    """Write all generated content to an Excel workbook organised by condition group."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    # Group results by condition group
    groups = {}
    for item in all_results:
        g = item["group"]
        groups.setdefault(g, []).append(item)

    header_fill = PatternFill("solid", fgColor="1a6b45")
    header_font = Font(bold=True, color="FFFFFF")
    day_fill    = PatternFill("solid", fgColor="d4edda")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )

    for group_name, items in groups.items():
        ws = wb.create_sheet(title=group_name[:31])

        # Header row
        headers = ["Day", "Topic", "Tip #", "Tip Content", "Source Hint", "Status"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border

        ws.row_dimensions[1].height = 22
        row = 2

        for item in sorted(items, key=lambda x: x["day_offset"]):
            tips = item.get("tips", [])
            if not tips:
                tips = [{"tip_number": 1, "tip": "(not generated)", "source_hint": ""}]

            for t in tips:
                # Alternate shading by day block
                fill = day_fill if item["day_offset"] % 10 < 5 else PatternFill()
                values = [
                    item["day_offset"],
                    item["topic"].replace("_", " ").title(),
                    t.get("tip_number", ""),
                    t.get("tip", ""),
                    t.get("source_hint", ""),
                    "raw — pending design",
                ]
                for col, val in enumerate(values, 1):
                    cell = ws.cell(row=row, column=col, value=val)
                    cell.border = thin_border
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                    if col <= 3:
                        cell.fill = fill
                row += 1

        # Column widths
        col_widths = [8, 22, 8, 70, 35, 22]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.freeze_panes = "A2"

    wb.save(output_path)
    print(f"  Excel saved → {output_path}")


# ---------------------------------------------------------------------------
# Main generation function
# ---------------------------------------------------------------------------

def generate_content(client_id: int, filter_group: str = None, filter_day: int = None,
                     dry_run: bool = False, no_db: bool = False,
                     output_dir: str = None) -> list:
    """
    Generate educational tips for all (or filtered) niche cases.
    Returns list of result dicts.
    """
    if output_dir is None:
        output_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "materials")
    os.makedirs(output_dir, exist_ok=True)

    cases = NICHE_CASES
    if filter_group:
        cases = [c for c in cases if c["group"] == filter_group]
    if filter_day:
        cases = [c for c in cases if c["day_offset"] == filter_day]

    if not cases:
        print("No niche cases match the given filters.")
        return []

    print(f"\n{'DRY RUN — ' if dry_run else ''}Generating content for {len(cases)} niche cases")
    print(f"  Client ID : {client_id}")
    print(f"  Output dir: {output_dir}\n")

    results = []
    db_session = None
    if not dry_run and not no_db:
        import database as db_module
        db_module.create_db_and_tables()
        db_session = db_module.SessionLocal()

    try:
        for i, niche in enumerate(cases, 1):
            label = f"[{i}/{len(cases)}] {niche['group']} / Day {niche['day_offset']} / {niche['topic']}"
            print(f"  {label}")

            if dry_run:
                results.append({**niche, "tips": [], "skipped": True})
                print(f"    → (dry run, skipping)")
                continue

            # Retrieve RAG chunks
            print(f"    → retrieving chunks...")
            try:
                chunks = _retrieve_chunks(niche["rag_query"], client_id)
            except Exception as e:
                print(f"    → retrieval error: {e} — continuing with empty context")
                chunks = ""

            # Generate tips
            print(f"    → generating tips via Ollama...")
            try:
                tips = _generate_tips(niche, chunks)
            except Exception as e:
                print(f"    → generation error: {e}")
                tips = []

            print(f"    → {len(tips)} tips generated")

            result = {**niche, "tips": tips}
            results.append(result)

            # Persist to DB
            if db_session and not no_db:
                try:
                    import database as db_module
                    db_module.upsert_content_material(
                        db_session,
                        condition_group=niche["group"],
                        condition_tags=niche["condition_tags"],
                        day_offset=niche["day_offset"],
                        topic=niche["topic"],
                        title=niche["title"],
                        raw_tips=tips,
                    )
                except Exception as e:
                    print(f"    → DB write error: {e}")

    finally:
        if db_session:
            db_session.close()

    # Write Excel
    if not dry_run and results:
        timestamp = datetime.now().strftime("%Y-%m-%d")
        excel_path = os.path.join(output_dir, f"content_library_{timestamp}.xlsx")
        try:
            _write_excel(results, excel_path)
        except Exception as e:
            print(f"  Excel write error: {e}")

    total = sum(len(r.get("tips", [])) for r in results)
    print(f"\nDone. {len(results)} niche cases, {total} tips generated.")
    return results


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Generate NutriBot educational content library")
    parser.add_argument("--client-id", type=int, default=4, help="Client ID for knowledge base (default: 4)")
    parser.add_argument("--group", type=str, help="Filter to one condition group (T2DM, HTN, CKD, Cardiac, PCOS, Dyslipidaemia, General)")
    parser.add_argument("--day", type=int, help="Filter to one day offset (3, 5, 7, 14, 21, 30)")
    parser.add_argument("--dry-run", action="store_true", help="Print what would be generated without calling LLM or writing DB")
    parser.add_argument("--no-db", action="store_true", help="Skip DB writes, only produce Excel")
    parser.add_argument("--output-dir", type=str, help="Output directory for Excel (default: materials/)")
    args = parser.parse_args()

    generate_content(
        client_id=args.client_id,
        filter_group=args.group,
        filter_day=args.day,
        dry_run=args.dry_run,
        no_db=args.no_db,
        output_dir=args.output_dir,
    )
