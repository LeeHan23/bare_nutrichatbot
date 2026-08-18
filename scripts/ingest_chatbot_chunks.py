"""
Ingest clinically-Approved rows from MyHeartCoach_Content_Registry.xlsx's
Chatbot_Chunks tab into the base_knowledge PGVector collection.

These are pre-written, clinically-approved Q&A grounding chunks (not raw
PDF text) — tagged trust_tier="clinical_approved" so TopicBoostedRetriever
ranks them above raw PDF chunks at equal topic overlap (vector_store.py).

Governance rule (workbook README tab): only Status == "Approved" rows ever
ship. Idempotent by Chunk ID — safe to re-run every time the client resends
the workbook; an edited Approved chunk's wording simply replaces the old
embedding, nothing accumulates.

    python scripts/ingest_chatbot_chunks.py MyHeartCoach_Content_Registry.xlsx
    python scripts/ingest_chatbot_chunks.py MyHeartCoach_Content_Registry.xlsx --dry-run
"""
import argparse
import os
import sys

_PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, _PROJECT_ROOT)

import openpyxl
from sqlalchemy import create_engine, text
from sqlalchemy.orm import sessionmaker
from dotenv import load_dotenv
from langchain_community.vectorstores import PGVector
from langchain_core.documents import Document

from taxonomy import COMPONENT_LABELS
from vector_store import get_connection_string

load_dotenv(os.path.join(_PROJECT_ROOT, ".env"))

PGVECTOR_URL = os.getenv("PGVECTOR_URL")
if not PGVECTOR_URL:
    raise SystemExit("PGVECTOR_URL not set")

COLLECTION_NAME = "base_knowledge"
_LABEL_TO_SLUG = {label.lower(): slug for slug, label in COMPONENT_LABELS.items()}


def _sheet_rows(ws):
    """Yield each data row (after the header) as a dict of header -> value."""
    rows = ws.iter_rows(min_row=1, max_row=ws.max_row)
    header = [c.value for c in next(rows)]
    for row in rows:
        values = [c.value for c in row]
        if not any(v is not None and str(v).strip() for v in values):
            continue
        yield dict(zip(header, values))


def build_content_id_to_component(wb) -> dict:
    """Content_Registry: Content ID -> component slug, best-effort. A row
    whose 'Components' text doesn't match a known label is left unmapped
    (the resulting chunk gets no doc_components tag — treated as legacy/
    unscoped by TopicBoostedRetriever, i.e. always eligible, never wrongly
    excluded)."""
    mapping = {}
    for row in _sheet_rows(wb["Content_Registry"]):
        content_id = row.get("Content ID")
        label = (row.get("Components") or "").strip()
        if not content_id or not label:
            continue
        slug = _LABEL_TO_SLUG.get(label.lower())
        if slug:
            mapping[content_id] = slug
        else:
            print(f"  [WARN] Content_Registry row {content_id}: unrecognized Component '{label}' — leaving untagged")
    return mapping


def build_documents(wb) -> list[Document]:
    content_id_to_component = build_content_id_to_component(wb)
    docs = []
    skipped = 0
    for row in _sheet_rows(wb["Chatbot_Chunks"]):
        if (row.get("Status") or "").strip() != "Approved":
            skipped += 1
            continue
        chunk_id = row.get("Chunk ID")
        answer = (row.get("Approved Answer Chunk") or "").strip()
        if not chunk_id or not answer:
            continue

        linked_id = row.get("Linked Content ID")
        component = content_id_to_component.get(linked_id)

        metadata = {
            "source": "chatbot_chunks",
            "chunk_id": chunk_id,
            "linked_content_id": linked_id,
            "chunk_type": row.get("Chunk Type"),
            "condition_tag": row.get("Condition Tag"),
            "persona_tag": row.get("Persona Tag"),
            "urgency_level": row.get("Urgency Level"),
            "follow_up_prompt": row.get("Follow-up Prompt"),
            "escalation_rule_ref": row.get("Escalation Rule Ref"),
            "trust_tier": "clinical_approved",
        }
        if component:
            metadata["doc_components"] = [component]

        question_intent = row.get("Question Intent") or ""
        page_content = f"{question_intent}\n{answer}" if question_intent else answer
        docs.append(Document(page_content=page_content, metadata=metadata))

    print(f"Chatbot_Chunks: {len(docs)} Approved row(s), {skipped} skipped (not Approved)")
    return docs


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", help="Path to MyHeartCoach_Content_Registry.xlsx")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    wb = openpyxl.load_workbook(args.workbook, data_only=True)
    docs = build_documents(wb)
    if not docs:
        print("No Approved Chatbot_Chunks rows found — nothing to do.")
        return

    engine = create_engine(PGVECTOR_URL)
    Session = sessionmaker(bind=engine)
    session = Session()
    collection_uuid = session.execute(
        text("SELECT uuid FROM langchain_pg_collection WHERE name = :name"),
        {"name": COLLECTION_NAME},
    ).scalar()

    for doc in docs:
        preview = doc.page_content[:80].replace("\n", " ")
        print(f"  {doc.metadata['chunk_id']} [{doc.metadata.get('doc_components', 'untagged')}]: {preview}...")

    if args.dry_run:
        print(f"\nDRY RUN — would upsert {len(docs)} chunk(s) into '{COLLECTION_NAME}'")
        session.close()
        return

    # Idempotent: delete any existing embedding for each chunk_id before
    # re-inserting, so re-running against a re-supplied workbook never
    # duplicates and always reflects the latest Approved wording.
    if collection_uuid:
        chunk_ids = [d.metadata["chunk_id"] for d in docs]
        result = session.execute(
            text(
                "DELETE FROM langchain_pg_embedding "
                "WHERE collection_id = :coll AND cmetadata->>'chunk_id' = ANY(:ids)"
            ),
            {"coll": collection_uuid, "ids": chunk_ids},
        )
        session.commit()
        if result.rowcount:
            print(f"Removed {result.rowcount} existing embedding(s) for re-ingested chunk_id(s)")
    session.close()

    from embeddings import get_embedding_function

    PGVector.from_documents(
        documents=docs,
        embedding=get_embedding_function(),
        collection_name=COLLECTION_NAME,
        connection_string=get_connection_string(),
        pre_delete_collection=False,
        use_jsonb=True,
    )
    print(f"\nUpserted {len(docs)} chunk(s) into '{COLLECTION_NAME}'")


if __name__ == "__main__":
    main()
